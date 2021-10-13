from bs4 import BeautifulSoup # Manipulacion de codigo html
import pandas as pd # Manipulacion de dataframe
import requests # Peticiones http
from openpyxl import load_workbook # Manipulacion de archivos xlsx
from openpyxl.styles.fills import PatternFill # Manipulacion de archivos xlsx
import datetime # Libreria para manipular fechas

# Carga y lectura del archivo excel

archivo_excel = pd.read_excel('inventario.xlsx')
wb = load_workbook(filename='inventario.xlsx')
sheet= wb.active

# Obtengo el tamaño del excel (cant. de filas y columnas)

cantidad_de_filas,ultima_columna = archivo_excel.shape

# Defino variables con valores iniciales

fila_referencia=2
cantidadLinkcaidos , cantidadLinkarriba= 0 , 0

# Defino nuevas Columnas

columna_estado_producto=ultima_columna+1
columna_fecha=ultima_columna+2

# Escribo columnas 

mycell= sheet.cell(row=1, column=columna_estado_producto)
mycell.value = "Estado producto"
mycell= sheet.cell(row=1, column=ultima_columna+2)
mycell.value = "Hora Actualizacion"

# Esto configura para ponerle color a las celdas

color_producto_0= PatternFill(fgColor="FF0000",
                   fill_type='solid')
color_producto_1= PatternFill(fgColor="77D970",
                   fill_type='solid')

# Itero sobre todas las columnas

for i in range(0, cantidad_de_filas):

    url_obtenida = archivo_excel.iloc[i, 1]

    # Guardamos en variable page contenido del sitio

    page = requests.get(url_obtenida)
    soup = BeautifulSoup(page.content, 'html.parser')

    # Convertimos a STR el contenido del HTML

    contenidoStr = str(soup.currentTag)

    # Buscar el texto 'Producto no encontrado' dentro de la variable contenidoStr

    buscarTexto = contenidoStr.find('Producto no encontrado')

    # Si la variable buscarTexto es mayor a 0, significa que no se encuentra (0),
    # de lo contrario, mostrará 'Producto encontrado' (1)

    if buscarTexto > 0:
        celda_producto= sheet.cell(row=fila_referencia, column=columna_estado_producto)
        celda_producto.value = 0
        celda_producto.fill = color_producto_0
        cantidadLinkcaidos +=1
    else:
        celda_producto= sheet.cell(row=fila_referencia, column=columna_estado_producto)
        celda_producto.value = 1
        celda_producto.fill = color_producto_1
        cantidadLinkarriba += 1

    celda_fecha= sheet.cell(row=fila_referencia, column=columna_fecha)
    celda_fecha.value= datetime.datetime.now()
    fila_referencia += 1

# Guardo el archivo con los cambios

wb.save(filename="inventario.xlsx")